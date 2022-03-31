from email.utils import collapse_rfc2231_value
from tkinter import dialog
from kivy.lang import Builder
from kivy.properties import StringProperty, NumericProperty
from kivy.uix.screenmanager import Screen
from kivymd.icon_definitions import md_icons
from kivymd.app import MDApp
from kivymd.uix.list import OneLineIconListItem
from kivy.uix.textinput import TextInput
from kivymd.uix.textfield import MDTextField
from kivymd.uix.button import MDRaisedButton
from kivy.core.audio import SoundLoader
from random import randrange
from kivy.config import Config
from kivy.properties import ObjectProperty
import sys
from kivy.properties import ObjectProperty, NumericProperty, StringProperty, \
    BooleanProperty, DictProperty, OptionProperty, ListProperty, ColorProperty
from kivy.logger import Logger
from kivy.graphics import Color, BorderImage, Canvas
from kivy.uix.textinput import TextInput

#importei daqui
from kivymd.uix.menu import MDDropdownMenu
from kivy.metrics import dp

#flatbutton e dialog
from kivymd.uix.button import MDFlatButton
from kivymd.uix.dialog import MDDialog
from kivymd.uix.list import OneLineAvatarIconListItem
from kivymd.uix.list import OneLineAvatarListItem
from kivy.uix.boxlayout import BoxLayout
from kivymd.uix.snackbar import Snackbar
from kivy.core.window import Window
from kivymd.uix.card import MDSeparator

#imports do excell

import kivy 
from kivy.app import App   
import os
from kivy.utils import platform
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

from kivy.uix.screenmanager import ScreenManager, Screen

kivy.require('2.0.0')
now1 = datetime.now()
dataAtual=now1.strftime("%d-%m-%Y %Hh%M")
dirdi=os.path.join(os.path.expandvars("%userprofile%"),"Documents\\Pap appy")

def log(pop):
    with open(dirdi+"Log.txt", 'a') as f:
        print(str(datetime.now())+" "+pop, file=f)
    print(pop)

if platform == 'android':

    import android

    from android.permissions import request_permissions, Permission
    #request_permissions([Permission.WRITE_EXTERNAL_STORAGE, Permission.READ_EXTERNAL_STORAGE])

    from android.storage import primary_external_storage_path

    #criação e checagem de diretorios
    dirPrincipal = os.path.join(r"/storage/emulated/0/Documents/Pap appy")
    dirDadosSalvos = os.path.join(r"/storage/emulated/0/Documents/Pap appy/Dados Salvos")
    
    try:
        os.makedirs(dirPrincipal)
        log("debug: O Diretório " + dirPrincipal +  " foi Criado pois ele não existe!")
    except FileExistsError:
        log("debug: O Diretório " + dirPrincipal +  " não foi criado pois ele já existe!")



    try:
        os.makedirs(dirDadosSalvos)
        log("debug: O Diretório " + dirDadosSalvos +  " foi Criado pois ele não existe!")
    except FileExistsError:
        log("debug: O Diretório " + dirDadosSalvos +  " não foi criado pois ele já existe!")
        
    #checagem do banco de dados raiz

    try:
        dirDoExcellSecundario = os.path.join(r"/storage/emulated/0/Documents/Pap appy/locais.xlsx")
        dirDoExcell = os.path.join(r"/storage/emulated/0/Documents/Pap appy/ticbk.xlsx")
        wb = openpyxl.load_workbook(dirDoExcell)
        wd= openpyxl.load_workbook(dirDoExcellSecundario)
    except FileNotFoundError as not_found:
        dirDoExcellSecundario = os.path.join(r"locais.xlsx")
        dirDoExcell = os.path.join(r"pap.xlsx")
        wb = openpyxl.load_workbook(dirDoExcell)
        wd= openpyxl.load_workbook(dirDoExcellSecundario)

    dirExcellDadosSalvos=os.path.join(r"/storage/emulated/0/Documents/Pap appy/Dados salvos/Dados salvos "+dataAtual+".xlsx")
    log("debug: o diretorio do excell é: "+dirDoExcell)
    
else:
    #Criação e checagem de diretorios
    dirPrincipal = os.path.join(os.path.expandvars("%userprofile%"),"Documents\\Pap appy")
    dirDadosSalvos = os.path.join(os.path.expandvars("%userprofile%"),"Documents\\Pap appy\\Dados Salvos")
   
    
    try:
        os.makedirs(dirPrincipal)
        log("debug: O Diretório " + dirPrincipal +  " foi Criado pois ele não existe!")
    except FileExistsError:
        log("debug: O Diretório " + dirPrincipal +  " não foi criado pois ele já existe!")



    try:
        os.makedirs(dirDadosSalvos)
        log("debug: O Diretório " + dirDadosSalvos +  " foi Criado pois ele não existe!")
    except FileExistsError:
        log("debug: O Diretório " + dirDadosSalvos +  " não foi criado pois ele já existe!")

    #Checagem do banco de dados raiz
    try:
        dirDoExcellSecundario = os.path.join(os.path.expandvars("%userprofile%"),"Documents\\Pap appy\\locais.xlsx")
        dirDoExcell = os.path.join(os.path.expandvars("%userprofile%"),"Documents\\Pap appy\\pap.xlsx")
        wd=openpyxl.load_workbook(dirDoExcellSecundario)
        wb = openpyxl.load_workbook(dirDoExcell)
        log("Debug: estamos utilizando o tic do docs,"+" o diretorio do excell é: "+dirDoExcell)

         
    except FileNotFoundError as not_found:
        dirDoExcellSecundario = os.path.join(os.path.expandvars("%userprofile%"),"Downloads\\vac.xlsx")
        dirDoExcell = os.path.join(os.path.expandvars("%userprofile%"),"Downloads\\vac2.xlsx") 
        wd=openpyxl.load_workbook(dirDoExcellSecundario)
        wb = openpyxl.load_workbook(dirDoExcell)
        log("debug: estamos utilizando o tic da raiz, o diretorio do excell é: "+ dirDoExcell)
    dirExcellDadosSalvos=os.path.join(os.path.expandvars("%userprofile%"),"Documents\\Pap appy\\Dados salvos\\Dados salvos "+dataAtual+".xlsx")


#configurações do excell

wz=wd.active
ws = wb.active

from openpyxl import Workbook
book = Workbook()
sheet = book.active
headers = ['CPF','Nome','Ultima Vacina','Vacinas Tomadas', '', '','Data de salvamento']
sheet.append(headers)      


#cores do app
colors = {
    "Red": {
        "50": "FFEBEE",
        "100": "FFCDD2",
        "200": "EF9A9A",
        "300": "E57373",
        "400": "EF5350",
        "500": "F44336",
        "600": "E53935",
        "700": "D32F2F",
        "800": "C62828",
        "900": "B71C1C",
        "A100": "FF8A80",
        "A200": "FF5252",
        "A400": "FF1744",
        "A700": "D50000",
    },
    "Pink": {
        "50": "FCE4EC",
        "100": "F8BBD0",
        "200": "F48FB1",
        "300": "F06292",
        "400": "EC407A",
        "500": "E91E63",
        "600": "D81B60",
        "700": "C2185B",
        "800": "AD1457",
        "900": "880E4F",
        "A100": "FF80AB",
        "A200": "FF4081",
        "A400": "F50057",
        "A700": "C51162",
    },
    "Purple": {
        "50": "F3E5F5",
        "100": "E1BEE7",
        "200": "CE93D8",
        "300": "BA68C8",
        "400": "AB47BC",
        "500": "9C27B0",
        "600": "8E24AA",
        "700": "7B1FA2",
        "800": "6A1B9A",
        "900": "4A148C",
        "A100": "EA80FC",
        "A200": "E040FB",
        "A400": "D500F9",
        "A700": "AA00FF",
    },
    "DeepPurple": {
        "50": "EDE7F6",
        "100": "D1C4E9",
        "200": "B39DDB",
        "300": "9575CD",
        "400": "7E57C2",
        "500": "673AB7",
        "600": "5E35B1",
        "700": "512DA8",
        "800": "4527A0",
        "900": "311B92",
        "A100": "B388FF",
        "A200": "7C4DFF",
        "A400": "651FFF",
        "A700": "6200EA",
    },
    "Indigo": {
        "50": "E8EAF6",
        "100": "C5CAE9",
        "200": "9FA8DA",
        "300": "7986CB",
        "400": "5C6BC0",
        "500": "3F51B5",
        "600": "3949AB",
        "700": "303F9F",
        "800": "283593",
        "900": "1A237E",
        "A100": "8C9EFF",
        "A200": "536DFE",
        "A400": "3D5AFE",
        "A700": "304FFE",
    },
    "Blue": {
        "50": "E3F2FD",
        "100": "BBDEFB",
        "200": "90CAF9",
        "300": "64B5F6",
        "400": "42A5F5",
        "500": "2196F3",
        "600": "1E88E5",
        "700": "1976D2",
        "800": "1565C0",
        "900": "0D47A1",
        "A100": "82B1FF",
        "A200": "448AFF",
        "A400": "2979FF",
        "A700": "2962FF",
    },
    "LightBlue": {
        "50": "E1F5FE",
        "100": "B3E5FC",
        "200": "81D4FA",
        "300": "4FC3F7",
        "400": "29B6F6",
        "500": "03A9F4",
        "600": "039BE5",
        "700": "0288D1",
        "800": "0277BD",
        "900": "01579B",
        "A100": "80D8FF",
        "A200": "40C4FF",
        "A400": "00B0FF",
        "A700": "0091EA",
    },
    "Cyan": {
        "50": "E0F7FA",
        "100": "B2EBF2",
        "200": "80DEEA",
        "300": "4DD0E1",
        "400": "26C6DA",
        "500": "00BCD4",
        "600": "00ACC1",
        "700": "0097A7",
        "800": "00838F",
        "900": "006064",
        "A100": "84FFFF",
        "A200": "18FFFF",
        "A400": "00E5FF",
        "A700": "00B8D4",
    },
    "Teal": {
        "50": "E0F2F1",
        "100": "B2DFDB",
        "200": "80CBC4",
        "300": "4DB6AC",
        "400": "26A69A",
        "500": "009688",
        "600": "00897B",
        "700": "00796B",
        "800": "00695C",
        "900": "004D40",
        "A100": "A7FFEB",
        "A200": "64FFDA",
        "A400": "1DE9B6",
        "A700": "00BFA5",
    },
    "Green": {
        "50": "E8F5E9",
        "100": "C8E6C9",
        "200": "A5D6A7",
        "300": "81C784",
        "400": "66BB6A",
        "500": "4CAF50",
        "600": "43A047",
        "700": "388E3C",
        "800": "2E7D32",
        "900": "1B5E20",
        "A100": "B9F6CA",
        "A200": "69F0AE",
        "A400": "00E676",
        "A700": "00C853",
    },
    "LightGreen": {
        "50": "F1F8E9",
        "100": "DCEDC8",
        "200": "C5E1A5",
        "300": "AED581",
        "400": "9CCC65",
        "500": "8BC34A",
        "600": "7CB342",
        "700": "689F38",
        "800": "558B2F",
        "900": "33691E",
        "A100": "CCFF90",
        "A200": "B2FF59",
        "A400": "76FF03",
        "A700": "64DD17",
    },
    "Lime": {
        "50": "F9FBE7",
        "100": "F0F4C3",
        "200": "E6EE9C",
        "300": "DCE775",
        "400": "D4E157",
        "500": "CDDC39",
        "600": "C0CA33",
        "700": "AFB42B",
        "800": "9E9D24",
        "900": "827717",
        "A100": "F4FF81",
        "A200": "EEFF41",
        "A400": "C6FF00",
        "A700": "AEEA00",
    },
    "Yellow": {
        "50": "FFFDE7",
        "100": "FFF9C4",
        "200": "FFF59D",
        "300": "FFF176",
        "400": "FFEE58",
        "500": "FFEB3B",
        "600": "FDD835",
        "700": "FBC02D",
        "800": "F9A825",
        "900": "F57F17",
        "A100": "FFFF8D",
        "A200": "FFFF00",
        "A400": "FFEA00",
        "A700": "FFD600",
    },
    "Amber": {
        "50": "FFF8E1",
        "100": "FFECB3",
        "200": "FFE082",
        "300": "FFD54F",
        "400": "FFCA28",
        "500": "FFC107",
        "600": "FFB300",
        "700": "FFA000",
        "800": "FF8F00",
        "900": "FF6F00",
        "A100": "FFE57F",
        "A200": "FFD740",
        "A400": "FFC400",
        "A700": "FFAB00",
    },
    "Orange": {
        "50": "FFF3E0",
        "100": "FFE0B2",
        "200": "FFCC80",
        "300": "FFB74D",
        "400": "FFA726",
        "500": "FF9800",
        "600": "FB8C00",
        "700": "F57C00",
        "800": "EF6C00",
        "900": "E65100",
        "A100": "FFD180",
        "A200": "FFAB40",
        "A400": "FF9100",
        "A700": "FF6D00",
    },
    "DeepOrange": {
        "50": "FBE9E7",
        "100": "FFCCBC",
        "200": "FFAB91",
        "300": "FF8A65",
        "400": "FF7043",
        "500": "FF5722",
        "600": "F4511E",
        "700": "E64A19",
        "800": "D84315",
        "900": "BF360C",
        "A100": "FF9E80",
        "A200": "FF6E40",
        "A400": "FF3D00",
        "A700": "DD2C00",
    },
    "Brown": {
        "50": "EFEBE9",
        "100": "D7CCC8",
        "200": "BCAAA4",
        "300": "A1887F",
        "400": "8D6E63",
        "500": "795548",
        "600": "6D4C41",
        "700": "5D4037",
        "800": "4E342E",
        "900": "3E2723",
        "A100": "000000",
        "A200": "000000",
        "A400": "000000",
        "A700": "000000",
    },
    "Gray": {
        "50": "FAFAFA",
        "100": "F5F5F5",
        "200": "EEEEEE",
        "300": "E0E0E0",
        "400": "BDBDBD",
        "500": "9E9E9E",
        "600": "757575",
        "700": "616161",
        "800": "424242",
        "900": "212121",
        "A100": "000000",
        "A200": "000000",
        "A400": "000000",
        "A700": "000000",
    },
    "BlueGray": {
        "50": "ECEFF1",
        "100": "CFD8DC",
        "200": "B0BEC5",
        "300": "90A4AE",
        "400": "78909C",
        "500": "607D8B",
        "600": "546E7A",
        "700": "455A64",
        "800": "37474F",
        "900": "263238",
        "A100": "000000",
        "A200": "000000",
        "A400": "000000",
        "A700": "000000",
    },
    "Light": {
        "StatusBar": "E0E0E0",
        "AppBar": "F5F5F5",
        "Background": "FAFAFA",
        "CardsDialogs": "FFFFFF",
        "FlatButtonDown": "cccccc",
    },
    "Dark": {
        "StatusBar": "000000",
        "AppBar": "1f1f1f",
        "Background": "121212",
        "CardsDialogs": "212121",
        "FlatButtonDown": "999999",
    },
}

#layout do app
Builder.load_string(
    '''
#:import images_path kivymd.images_path
#:import Snackbar kivymd.uix.snackbar.Snackbar
     
<Item>

    IconLeftWidget:
        icon: root.icon 
<IconListItem>
    IconLeftWidget:
        icon: root.icon
#
#
# 
# 
# --------------------------Layout Da caixa de confirmaçã--------------------------------------
#
#
#
<Contentx>
    orientation: "vertical"
    spacing: "12dp"
    size_hint_y: None
    height: "360dp"
    padding:"12dp"

    MDLabel:
        text: "CPF: "+app.namez
        font_style:"Subtitle2"
    MDTextField:
        mode: "rectangle"
        hint_text: ""
        text: app.dialogCpf
        line_color_normal:"#398c36"
        text_color_normal:"#398c36"
        disabled:True
    MDLabel:
        text: "Ultima Vacina:"
        font_style:"Subtitle2"
    MDTextField:
        mode: "rectangle"
        hint_text: ""
        text: app.dialogTurno
        line_color_normal:"#398c36"
        text_color_normal:"#398c36"
        disabled:True
    MDLabel:
        text: "Doses Tomadas:"
        font_style:"Subtitle2"
    MDTextField:
        mode: "rectangle"
        markup: True
        font_style: 'H1'
        hint_text: ""
        text: app.dialogVacinas
        line_color_normal:"#398c36"
        text_color_normal:"#398c36"
        disabled:True

<CustomOneLineIconListItem>
    
    IconLeftWidget:
        icon: root.icon  
#
#
#
#
#
#--------------------------Layout Principal--------------------------------------
#
#
#
<PreviousMDIcons>
    MDBoxLayout:
        orientation: 'vertical'
        spacing: dp(0)
        padding: dp(0)
        MDBoxLayout:
            adaptive_height: True
    
            MDToolbar:
                title: ""

        MDBoxLayout:
            spacing: dp(0)
            padding: dp(60)
            orientation: 'vertical'

            Image:
                size_hint_y: None
                id: bg_image
                source: "fpsicon.png"
                pos_hint: {'center_x': .5, 'center_y': .5}
                width: 100
                allow_stretch: True
    
            MDLabel:
                text: "Lista de Vacinados"
                halign:"center"
                font_style:"H6"

            MDLabel:
                text: "Insira as informações abaixo para realizar uma consulta"
                halign:"center"     

            MDTextField:
                id:campoCpf
                mode: "rectangle" 
                width: "2dp"
                text_color_normal:app.theme_cls.primary_color
                hint_text_color_normal:app.theme_cls.accent_color
                current_hint_text_color:[0.23529411764705882, 0.2549019607843137, 0.25882352941176473, 1.0] #mudou borda e letra 
                text_color:app.theme_cls.accent_color  #mudou cor do texto selecionado pra azul
                hint_text: "CPF:"                           
                helper_text: ""
                helper_text_mode: "on_focus"
                on_focus: app.campo_Cpf_Selecionado()
                on_text_validate: app.campo_Cpf_Selecionado()

            MDLabel:
                id:textoNome
                theme_text_color: "Custom"
                text: "Nome:"
                font_style:"Subtitle2"
                text_color:"#09101D"
                
            MDTextField:
                id: campoNome
                active_line:False
                icon_right: "card-account-details-outline"
                mode: "rectangle"
                font_size:'15sp'
                #icon_right: "card-account-details-outline" 
                disabled:True                       
                helper_text: "você pode digitar se preferir"
                helper_text_mode: "on_focus"
                text_color_normal:app.theme_cls.primary_color
                hint_text_color_normal:app.theme_cls.accent_color
                current_hint_text_color:[0.23529411764705882, 0.2549019607843137, 0.25882352941176473, 1.0] #mudou borda e letra 
                text_color:app.theme_cls.accent_color  #mudou cor do texto selecionado pra azul
                on_focus: app.campo_nome_Selecionado()
                
            MDLabel:
                id:textoNome
                theme_text_color: "Custom"
                text: "Doses Tomadas:"
                font_style:"Subtitle2"
                text_color:"#09101D"
    
            MDTextField:
                id: campoVacinas 
                active_line:False
                icon_right: "needle"
                mode: "rectangle"
                font_size:'15sp'
                #icon_right: "database-search" 
                disabled:True                       
                helper_text: "você pode digitar se preferir"
                helper_text_mode: "on_focus"
                 
                text_color_normal:app.theme_cls.primary_color
                hint_text_color_normal:app.theme_cls.accent_color
                current_hint_text_color:[0.23529411764705882, 0.2549019607843137, 0.25882352941176473, 1.0] #mudou borda e letra 
                text_color:app.theme_cls.accent_color  #mudou cor do texto selecionado pra azul
                on_focus: app.campo_nome_Selecionado()
                
            MDLabel:
                id:textoAjudante
                text: "Its Search Time!!"
                font_style:"Subtitle2"
                theme_text_color: "Custom"
                text_color:"#B95000"
                
            RecycleView:
                id: rv
                key_viewclass: 'viewclass'
                key_size: 'height'
                RecycleBoxLayout:
                    padding: dp(10)
                    default_size: None, dp(48)
                    default_size_hint: 1, None
                    size_hint_y: None
                    height: self.minimum_height
                    orientation: 'vertical' 

        MDFloatingActionButton:
            icon: "database-search"
            on_release: app.checarDados(campoCpf.text)
            md_bg_color: app.theme_cls.primary_color
            pos_hint: {'center_x': .5, 'center_y': .5}

'''
)

#classe do turno abaixo:
class IconListItem(OneLineIconListItem):
    icon = StringProperty()

class Contentx(BoxLayout):
    pass


#classe do dialog
class Item(OneLineAvatarListItem):
    icon = StringProperty()
    divider = None
    source = StringProperty()

class CustomSnackbar(Snackbar):
    text = StringProperty(None)
    icon = StringProperty(None)
    font_size = NumericProperty("15sp")


class CustomDialog(MDDialog):
    title = StringProperty()
    icon = StringProperty()

class CapitalInput(MDTextField):
    def insert_text(self, substring, from_undo=False):
        s = substring.upper()
        return super(CapitalInput, self).insert_text(s, from_undo=from_undo)

class CustomOneLineIconListItem(OneLineIconListItem):
    icon = StringProperty()
    
# Declare both screens
class MenuScreen(Screen):
    pass

class SettingsScreen(Screen):
    pass
class PreviousMDIcons(Screen):
    pass
        


class MainApp(MDApp):
    dialog = None
    dialog2=None
    dialog3=None
    namez=StringProperty("mat")
    first_namez=StringProperty("mat")
    dialogCpf=StringProperty("")
    dialogTurno=StringProperty("")
    dialogVacinas=StringProperty("")
    dialogUltimaVacina=StringProperty("")

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.screen = PreviousMDIcons() 

#
#
#-----------------checadores de dados------------------
#
#

    def campo_Cpf_Selecionado(self):
        if self.screen.ids.campoCpf.text=="":
           print(self.screen.ids.campoCpf.line_color_normal)
           print(self.screen.ids.campoCpf.text,"hmm")
  
        else:
           print(self.screen.ids.campoCpf.current_hint_text_color)
           print(self.screen.ids.campoCpf.text,"hmm")
           
           self.checarDados2(self.screen.ids.campoCpf.text)           

    def campo_nome_Selecionado(self):
        self.screen.ids.campoVacinas.icon_right="chevron-up"
        if self.screen.ids.campoVacinas.text=="":
            print(self.screen.ids.campoVacinas.current_hint_text_color)
        else:  
            print("po")

    def checar_cpf(self):
        if not self.screen.ids.campoCpf.text:
            if self.logado==True:
                self.screen.ids.textoAjudante.text="Você precisa preencher o cpf"
            else:
                self.screen.ids.textoAjudante.text="Você precisa Logar em uma conta para continuar"
            
            self.screen.ids.textoAjudante.text_color = "#B95000"    
            self.screen.ids.campoCpf.line_color_normal="#B95000"
            self.screen.ids.campoCpf.hint_text_color_normal="#09101D"
            self.screen.ids.campoCpf.line_color_focus="#18a0fb"
            self.screen.ids.campoCpf.hint_text_color_focus="#18a0fb"

        else:
            if self.logado==True:
                self.screen.ids.textoAjudante.text="cpf foi corretamente selecionada"
            else:
                self.screen.ids.textoAjudante.text="cpf foi corretamente selecionada"
            self.screen.ids.textoAjudante.text_color = "#287D3C"
            self.screen.ids.campoCpf.line_color_focus="#18a0fb"
            self.screen.ids.campoCpf.line_color_normal="#287D3C"
            self.screen.ids.campoCpf.hint_text_color_normal="#09101D"
            self.screen.ids.campoCpf.hint_text_color_focus="#18a0fb"           

    def checarDados(self,cpf):
        self.dialogCpf=self.screen.ids.campoCpf.text
        self.dialogVacinas=self.screen.ids.campoVacinas.text
        self.dialogUltimaVacina="Rota "+str(self.rota)
        self.checar_cpf()
        if self.logado==True:
            if self.screen.ids.campoCpf.text=="":
                self.matriculafinder="Preencha os campos obrigatórios para poder salvar! 3 vazios"
            elif self.screen.ids.campoCpf.text in self.block:
                self.matriculafinder="este cpf já foi utilizada em um agendamento, peça a sua alteração na sala!"
                self.screen.ids.textoAjudante.text="este cpf já foi registrada no sistema!"
                self.screen.ids.textoAjudante.text_color="#DA1414"
                self.screen.ids.campoCpf.line_color_normal="#DA1414"
            else:
                if self.logado==True:
                    if self.cola1.__contains__(self.screen.ids.campoCpf.text):
                        self.wsid=1

                        self.domat(self.screen.ids.campoCpf.text)

                        self.show_confirmation_dialog()
                        self.screen.ids.textoAjudante.text_color = "#287D3C"
                        if self.screen.ids.campoVacinas.text=="4" or self.screen.ids.campoVacinas.text=="3":
                            self.screen.ids.textoAjudante.text="Um Resultado foi Encontrado, o cpf possui a quantidade de vacinas indicada para passar pela catraca!"
                        elif self.screen.ids.campoVacinas.text=="2" or self.screen.ids.campoVacinas.text=="1":  
                            self.screen.ids.textoAjudante.text_color = "#B95000"  
                            self.screen.ids.textoAjudante.text="Um Resultado foi Encontrado, o cpf não possui quantidade de vacinas indicada para passar pela catraca!"      
                    elif self.cola2.__contains__(self.screen.ids.campoCpf.text):
                        self.wsid=2

                        self.domat(self.screen.ids.campoCpf.text)

                        self.show_confirmation_dialog()
                        self.screen.ids.textoAjudante.text_color = "#287D3C"
                        if self.screen.ids.campoVacinas.text=="4" or self.screen.ids.campoVacinas.text=="3":
                            self.screen.ids.textoAjudante.text="Um Resultado foi Encontrado, o cpf possui a quantidade de vacinas indicada para passar pela catraca!"
                        elif self.screen.ids.campoVacinas.text=="2" or self.screen.ids.campoVacinas.text=="1":  
                            self.screen.ids.textoAjudante.text_color = "#B95000"  
                            self.screen.ids.textoAjudante.text="Um Resultado foi Encontrado, o cpf não possui quantidade de vacinas indicada para passar pela catraca!"      
                    else:
                        self.matriculafinder="o cpf digitada não existe no sistema!"
                        self.screen.ids.textoAjudante.text="Nenhum Resultado foi encontrado!"
                        self.screen.ids.textoAjudante.text_color ="#DA1414"  
                        self.screen.ids.campoCpf.line_color_normal="#DA1414"
                        print("colab enchendo e number")
                        print(self.matriculafinder)
                else:
                    self.screen.ids.textoAjudante.text_color="#DA1414"
                    self.screen.ids.textoAjudante.text="Você precisa Logar em uma conta para continuar"  
        else:
            self.screen.ids.textoAjudante.text_color="#DA1414"
            self.screen.ids.textoAjudante.text="Você precisa Logar em uma conta para continuar"  
    def checarDados2(self,cpf):
        self.dialogCpf=self.screen.ids.campoCpf.text
        self.dialogVacinas=self.screen.ids.campoVacinas.text
        self.dialogUltimaVacina="Rota "+str(self.rota)
        self.checar_cpf()
        
        if self.screen.ids.campoCpf.text=="":
            self.matriculafinder="Preencha os campos obrigatórios para poder salvar! 3 vazios"
        elif self.screen.ids.campoCpf.text in self.block:
            self.matriculafinder="este cpf já foi utilizada em um agendamento, peça a sua alteração na sala!"
            self.screen.ids.textoAjudante.text="este cpf já foi registrada no sistema!"
            self.screen.ids.textoAjudante.text_color="#DA1414"
            self.screen.ids.campoCpf.line_color_normal="#DA1414"
        else:

            if self.cola1.__contains__(self.screen.ids.campoCpf.text):
                self.wsid=1

                self.domat(self.screen.ids.campoCpf.text)
                

                self.screen.ids.textoAjudante.text_color = "#287D3C"
                if self.screen.ids.campoVacinas.text=="4" or self.screen.ids.campoVacinas.text=="3":
                    self.screen.ids.textoAjudante.text="Um Resultado foi Encontrado, o cpf possui a quantidade de vacinas indicada para passar pela catraca!"
                elif self.screen.ids.campoVacinas.text=="2" or self.screen.ids.campoVacinas.text=="1":    
                    self.screen.ids.textoAjudante.text_color = "#B95000"  
                    self.screen.ids.textoAjudante.text="Um Resultado foi Encontrado, o cpf não possui quantidade de vacinas indicada para passar pela catraca!"
            elif self.cola2.__contains__(self.screen.ids.campoCpf.text):
                self.wsid=2

                self.domat(self.screen.ids.campoCpf.text)
                self.screen.ids.textoAjudante.text_color = "#287D3C"
                if self.screen.ids.campoVacinas.text=="4" or self.screen.ids.campoVacinas.text=="3":
                    self.screen.ids.textoAjudante.text="Um Resultado foi Encontrado, o cpf possui a quantidade de vacinas indicada para passar pela catraca!"
                elif self.screen.ids.campoVacinas.text=="2" or self.screen.ids.campoVacinas.text=="1":  
                    self.screen.ids.textoAjudante.text_color = "#B95000"  
                    self.screen.ids.textoAjudante.text="Um Resultado foi Encontrado, o cpf não possui quantidade de vacinas indicada para passar pela catraca!"      
                    
            else:
                self.matriculafinder="o cpf digitada não existe no sistema!"

                self.screen.ids.textoAjudante.text="Nenhum Estudante encontrado!"  
                self.screen.ids.textoAjudante.text_color ="#DA1414"  
                self.screen.ids.campoCpf.line_color_normal="#DA1414"
                self.screen.ids.campoVacinas.text=""
                print("colab 1 :")
                print(self.matriculafinder)
#
#
#-----------------Salvar dados ------------------------
#
#
    def domat(self,cpf):
        if self.wsid==1:
            ws=wb.active
        else:
            ws=wd.active
        
        self.matriculafinder="Nenhuma cpf encontrada, verifique os dados e tente novamente!"
        encouter=False
        print("valour de  encouter é: ", str(encouter))
        
        log("debug: Lista de bloqueados - "+str(self.block))

        for i in range(1, ws.max_row + 1):

            if str(cpf) == str(ws.cell(i,1).value):
                    encouter=True
                    if encouter == True:
                        log("debug: uma matricula foi encontrada e uma caixa de dialogo foi acionada!")
                        log("debug: cpf encontrada e o valour de  encouter é: "+ str(encouter))
                    rox=(ws.cell(i,1).value)
                    self.namez=str(ws.cell(i,2).value)
                    self.dialogTurno=str(ws.cell(i,5).value)
                    self.dialogVacinas=str(ws.cell(i,7).value)
                    self.dialogUltimaVacina=str(ws.cell(i,3).value)
                    log("debug: Nome da aluna encontrada - "+str(self.namez))
                    self.screen.ids.campoNome.text=ws.cell(i,2).value
                    self.screen.ids.campoVacinas.text=str(ws.cell(i,7).value)

    def doThis(self,cpf):
        #request_permissions([Permission.WRITE_EXTERNAL_STORAGE, Permission.READ_EXTERNAL_STORAGE])
        self.matriculafinder="Nenhuma cpf encontrada, verifique os dados e tente novamente!"
        encouter=False
        log("debug: o valor de  encouter é: "+ str(encouter))

        for i in range(1, ws.max_row + 1):

            if str(cpf) == str(ws.cell(i,1).value):
                    encouter=True
                    log("debug: cpf encontrada e o valor de  encouter é: "+ str(encouter))
                    now = datetime.now()
                    dt=now.strftime("%d/%m/%Y %H:%M:%S")
                    if encouter == True:
                        log("debug: cpf encontrada e o valour de  encouter é: "+ str(encouter))
                    self.block.append(ws.cell(i,1).value)
                    log("debug: Lista de bloqueados - "+str(self.block))
                    rox=(ws.cell(i,1).value,ws.cell(i,2).value,ws.cell(i,5).value,ws.cell(i,7).value,"",dt)
                    
                    name=str(ws.cell(i,1).value)
                    self.first_name = name.rsplit(' ', 3)[0]
                    self.matriculafinder= f'As informações foram salvas e uma copia do protocolo foi enviada para o estudante!'
                    self.show(self.matriculafinder)
                    
                    rows = (
                        (rox[0],rox[1], rox[2], rox[3],rox[4],rox[5],),

                        )

                    for row in rows:
                        sheet.append(row)
                    
                    self.limpar()

                    book.save(dirExcellDadosSalvos)
                    
            elif str(cpf) != str(ws.cell(i,2).value):
                self.matriculafinder="Nenhuma cpf encontrada, verifique os dados e tente novamente!zzz"
                if encouter == True:
                    self.limpar()

#    
#
#-------------------------Snackbar--------------------------------
#
#
    def show(self,tt):
        self.snackbar = CustomSnackbar(
            text=tt,
            icon="information",
            snackbar_x="10dp",
            snackbar_y="10dp",
            bg_color="#4caf50",

            buttons=[MDFlatButton(text="[color=#FFFFFF]"+"OK"+"[/color]", on_release= self.close,text_color=self.theme_cls.primary_color,),]
            
        )
        self.snackbar.size_hint_x = (
            Window.width - (self.snackbar.snackbar_x * 2)
        ) / Window.width
        self.snackbar.open()

    def limpar(self):
        self.screen.ids.campoVacinas.text=""
        self.screen.ids.campoCpf.text=""
        self.screen.ids.campoNome.text=""
        self.screen.ids.textoAjudante.text="Insira um CPF para realizar uma pesquisa"
        self.screen.ids.textoAjudante.text_color = "#09101D"
        self.screen.ids.textoNome.text_color = "#09101D"
#
#    
#----------------------dialogbox-----------------------------------------
#
#
    def close(self, *args):
        self.snackbar.dismiss(self)

    def show_confirmation_dialog(self):
        if not self.dialog:
            self.dialog = CustomDialog(
                title="Confira os detalhes da vacina",
                icon="content-save",            
                type="custom",
                content_cls=Contentx(),
                buttons=[
                    MDRaisedButton(
                        text="Editar informações", md_bg_color=[.10, .10, .10, .10], text_color=[1,1,1,1],on_release= self.fecharDialogo
                    ),
                    MDRaisedButton(
                        text="Confirmar e salvar",md_bg_color=self.theme_cls.primary_color,text_color=[1,1,0,1], on_release= self.salvarDialogo
                    ),
                ],
            )
        self.dialog.open()
        


    def salvarDialogo(self, *args):
        self.dialog.dismiss(force=True)
        log("debug: cpf que foi salva - "+self.dialogCpf)
        cpf=self.dialogCpf
        self.dialogCpf=""
        self.dialogTurno=""
        self.dialogVacinas=""
        self.dialogUltimaVacina=""
        self.doThis(cpf)

    def fecharDialogo(self, *args):
        self.dialog.dismiss(force=True)
##
##-------------------Main configurations------------------------------------------------------
##
##           
    def build(self):
        self.theme_cls.colors = colors
        self.theme_cls.primary_palette = "Green"
        self.theme_cls.accent_palette = "Blue"
        self.theme_cls.secondary_palette ="Gray"
        return self.screen
    
    def on_start(self):
        self.wsid=1
        self.screen.ids.campoCpf.line_color_focus="#18a0fb"
        self.screen.ids.campoCpf.hint_text_color_focus="#18a0fb"
        self.cola1=[]
        self.cola2=[]
        self.colab=[]
        self.logado=True
        self.block=[]
        #self.screen.set_list_md_icons()
        self.rota=""
        self.first_name=""
        self.cpfselect=False
        for i in range(1, ws.max_row + 1):
            self.cola1.append(ws.cell(i,1).value)
        for i in range(1, wz.max_row + 1):
            self.cola2.append(wz.cell(i,1).value)
            
MainApp().run()

